from io import BytesIO

from openpyxl import load_workbook


def parse_xlsx(file_bytes: bytes) -> str:
    workbook = load_workbook(BytesIO(file_bytes), data_only=True)
    lines = []
    for sheet in workbook.worksheets:
        lines.append(f"# Sheet: {sheet.title}")
        for row in sheet.iter_rows(values_only=True):
            cells = [str(c) for c in row if c is not None]
            if cells:
                lines.append(" | ".join(cells))
    return "\n".join(lines).strip()
